import io
import csv
import os
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
import json
from model.inputs import ModelInputs, CountryConfig
from model.engine import calculate

app = FastAPI(title="WSS Scenarios Model API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/api/defaults")
def get_defaults():
    return ModelInputs().model_dump()


@app.get("/api/defaults/blank")
def get_blank():
    """Returns a blank template with zero values for a new country."""
    blank = ModelInputs(
        country_config=CountryConfig(country="", area="", currency="USD",
            provider1_name="Provider 1", provider2_name="Provider 2"),
    ).model_dump()
    # Zero out all data-specific fields but keep structure
    for section in ['macro', 'population', 'water_service', 'sanitation_service',
                    'water_costs', 'sanitation_costs', 'bau', 'technical']:
        if section in blank:
            for key, val in blank[section].items():
                if isinstance(val, (int, float)):
                    blank[section][key] = 0
                elif isinstance(val, list):
                    blank[section][key] = [0] * len(val)
    # Keep period defaults
    blank['period'] = {'model_start_year': 2020, 'forecast_end_year': 2050,
        'baseline_year': 2024, 'as_is_forecast_start': 2025,
        'as_is_forecast_length': 2, 'target1_year': 2030, 'target2_year': 2040}
    return blank


@app.get("/api/profiles")
def list_profiles():
    """List saved country profiles."""
    profiles_dir = os.path.join(os.path.dirname(__file__), "profiles")
    if not os.path.exists(profiles_dir):
        return []
    return [f.replace('.json', '') for f in os.listdir(profiles_dir) if f.endswith('.json')]


@app.get("/api/profiles/{name}")
def get_profile(name: str):
    """Load a saved country profile."""
    filepath = os.path.join(os.path.dirname(__file__), "profiles", f"{name}.json")
    if not os.path.exists(filepath):
        return {"error": "Profile not found"}
    with open(filepath) as f:
        return json.load(f)


@app.post("/api/profiles/{name}")
def save_profile(name: str, inputs: ModelInputs):
    """Save current inputs as a country profile."""
    profiles_dir = os.path.join(os.path.dirname(__file__), "profiles")
    os.makedirs(profiles_dir, exist_ok=True)
    filepath = os.path.join(profiles_dir, f"{name}.json")
    with open(filepath, 'w') as f:
        json.dump(inputs.model_dump(), f, indent=2)
    return {"status": "saved", "name": name}


@app.post("/api/calculate")
def run_calculation(inputs: ModelInputs):
    return calculate(inputs)


@app.post("/api/export/csv")
def export_csv(inputs: ModelInputs):
    result = calculate(inputs)

    years = result['years']
    total_hh = result['total_hh']
    ws = result['water_supply']
    san = result['sanitation']

    columns = [
        'Year', 'Total_HH',
        'WS_BAU_Serv1', 'WS_Target_Serv1', 'WS_Service_Gap',
        'WS_Investment_Need', 'WS_BAU_Investment', 'WS_Financing_Gap',
        'SAN_BAU_Serv1', 'SAN_Target_Serv1', 'SAN_Service_Gap',
        'SAN_Investment_Need', 'SAN_BAU_Investment', 'SAN_Financing_Gap',
    ]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)

    for i, year in enumerate(years):
        writer.writerow([
            year,
            total_hh[i],
            ws['bau_hh_serv'][0][i],
            ws['target_hh_serv'][0][i],
            ws['service_gap'][i],
            ws['investment_need'][i],
            ws['bau_investment'][i],
            ws['financing_gap'][i],
            san['bau_hh_serv'][0][i],
            san['target_hh_serv'][0][i],
            san['service_gap'][i],
            san['investment_need'][i],
            san['bau_investment'][i],
            san['financing_gap'][i],
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename="wss_results.csv"'},
    )


@app.post("/api/export/pptx")
def export_pptx(inputs: ModelInputs):
    from export_pptx import create_pptx
    result = calculate(inputs)
    output = create_pptx(result, inputs.model_dump())
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type='application/vnd.openxmlformats-officedocument.presentationml.presentation',
        headers={'Content-Disposition': 'attachment; filename="wss_scenarios.pptx"'},
    )


@app.post("/api/export/xlsx")
def export_xlsx(inputs: ModelInputs):
    from openpyxl import Workbook
    result = calculate(inputs)
    wb = Workbook()

    for sector_key, sector_name in [('water_supply', 'Water Supply'), ('sanitation', 'Sanitation')]:
        ws = wb.create_sheet(title=sector_name)
        sec = result[sector_key]
        headers = ['Year', 'Total HH', 'Target Serv1', 'BAU Serv1', 'Service Gap',
                   'Investment Need', 'BAU Investment', 'Financing Gap']
        ws.append(headers)
        for i, year in enumerate(result['years']):
            ws.append([
                year, result['total_hh'][i],
                sec['target_hh_serv'][0][i], sec['bau_hh_serv'][0][i], sec['service_gap'][i],
                sec['investment_need'][i], sec['bau_investment'][i], sec['financing_gap'][i],
            ])

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="wss_results.xlsx"'},
    )


@app.get("/api/health")
def health():
    return {"status": "ok"}


# Serve frontend static files
static_dir = os.path.join(os.path.dirname(__file__), "static")
if os.path.exists(static_dir):
    app.mount("/assets", StaticFiles(directory=os.path.join(static_dir, "assets")), name="assets")

    @app.get("/{full_path:path}")
    def serve_frontend(full_path: str):
        file_path = os.path.join(static_dir, full_path)
        if full_path and os.path.isfile(file_path):
            return FileResponse(file_path)
        return FileResponse(os.path.join(static_dir, "index.html"))
